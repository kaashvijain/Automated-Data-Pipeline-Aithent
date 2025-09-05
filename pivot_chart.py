import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

df = pd.read_csv('students_20250801.csv')

output_excel = 'students_pivot_charts.xlsx'
df.to_excel(output_excel, index=False)

wb = load_workbook(output_excel)
ws = wb.active

# chart info 
def add_subject_chart(subject, col_index, chart_pos):
    chart = BarChart()
    chart.title = f"{subject} Marks by Student"
    chart.y_axis.title = "Marks"
    chart.x_axis.title = "Student"
    data = Reference(ws, min_col=col_index, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, chart_pos)

# subject wise charts 
add_subject_chart('MATHS', 3, "H2")
add_subject_chart('ENGLISH', 4, "H20")
add_subject_chart('SCIENCE', 5, "H38")

wb.save(output_excel)

print("pivot charts created in 'students_pivot_charts.xlsx'")
